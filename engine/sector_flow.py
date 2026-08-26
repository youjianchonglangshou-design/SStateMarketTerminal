"""Daily US sector-flow collector for the HUD compass.

Primary source: State Street / SPDR daily NAV-history spreadsheets for the
11 Select Sector SPDR ETFs.  Each spreadsheet is public, keyless, and contains
Date, NAV, Shares Outstanding, and Total Net Assets back to inception.

Estimated daily primary-market flow:

    flow_usd = (shares_t - shares_t-1) * nav_t
    flow_pct = flow_usd / total_net_assets_t-1 * 100

This removes price-driven AUM changes from the flow estimate.  FLOW STRENGTH is
computed from the previous 20 valid trading-session flow_pct observations:

    z = (today_flow_pct - mean(previous_20_flow_pct)) / stdev(previous_20_flow_pct)
    strength = clamp(50 + 20*z, 0, 100)

The collector rejects stale or mixed-date source data instead of publishing an
old snapshot as if it were current.
"""

from __future__ import annotations

import argparse
import calendar
import json
import math
import statistics
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import requests

TW = timezone(timedelta(hours=8))
NAVHIST_BASE = "https://www.ssga.com/library-content/products/fund-data/etfs/us/navhist-us-en-{ticker}.xlsx"
MIN_HISTORY = 20
HISTORY_KEEP = 90

SECTORS: list[dict[str, str]] = [
    {"id": "TECH", "label": "科技", "name": "Information Technology", "ticker": "XLK"},
    {"id": "FIN", "label": "金融", "name": "Financials", "ticker": "XLF"},
    {"id": "HEALTH", "label": "醫療", "name": "Health Care", "ticker": "XLV"},
    {"id": "CD", "label": "非必需消費", "name": "Consumer Discretionary", "ticker": "XLY"},
    {"id": "CS", "label": "必需消費", "name": "Consumer Staples", "ticker": "XLP"},
    {"id": "COMM", "label": "通訊", "name": "Communication Services", "ticker": "XLC"},
    {"id": "IND", "label": "工業", "name": "Industrials", "ticker": "XLI"},
    {"id": "ENERGY", "label": "能源", "name": "Energy", "ticker": "XLE"},
    {"id": "UTIL", "label": "公用事業", "name": "Utilities", "ticker": "XLU"},
    {"id": "RE", "label": "房地產", "name": "Real Estate", "ticker": "XLRE"},
    {"id": "MAT", "label": "原物料", "name": "Materials", "ticker": "XLB"},
]


def _norm_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().lower().split())


def _num(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        raise ValueError("missing numeric value")
    if isinstance(value, (int, float)):
        out = float(value)
    else:
        s = str(value).strip().replace(",", "").replace("$", "").replace(" ", "")
        if not s or s in {"-", "—", "–"}:
            raise ValueError("missing numeric value")
        mult = 1.0
        if s[-1:].upper() in {"K", "M", "B", "T"}:
            unit = s[-1].upper()
            s = s[:-1]
            mult = {"K": 1e3, "M": 1e6, "B": 1e9, "T": 1e12}[unit]
        out = float(s) * mult
    if not math.isfinite(out):
        raise ValueError("non-finite numeric value")
    return out


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value or "").strip()
    if not s:
        raise ValueError("missing date")
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%b %d %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"unsupported date: {s}")


def _normalize_nav_row(raw: dict[str, Any]) -> dict[str, Any]:
    d = _parse_date(raw["date"])
    nav = _num(raw["nav"])
    shares_raw = _num(raw["shares_outstanding"])
    assets_raw = _num(raw["total_net_assets"])

    if nav <= 0 or shares_raw <= 0 or assets_raw <= 0:
        raise ValueError("NAV history row has non-positive values")

    # SSGA spreadsheet cells can be stored either as whole units or as values
    # displayed in millions.  Infer the unit combination from NAV * shares ~= AUM
    # instead of assuming one fixed workbook representation.
    candidates = []
    for share_scale in (1.0, 1_000_000.0):
        for asset_scale in (1.0, 1_000_000.0):
            shares = shares_raw * share_scale
            assets = assets_raw * asset_scale
            ratio = nav * shares / assets
            candidates.append((abs(ratio - 1.0), ratio, shares, assets))
    _, ratio, shares, assets = min(candidates, key=lambda x: x[0])
    if not (0.90 <= ratio <= 1.10):
        raise ValueError(
            f"NAV history unit check failed: best nav*shares/assets={ratio:.4f} "
            f"(nav={nav}, shares_raw={shares_raw}, assets_raw={assets_raw})"
        )
    return {
        "date": d.isoformat(),
        "nav": nav,
        "shares_outstanding": shares,
        "total_net_assets": assets,
    }


def parse_nav_history_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl unavailable for State Street NAV history: {exc}") from exc

    if not content or len(content) < 100:
        raise ValueError("State Street NAV history file is empty")
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"State Street NAV history is not a readable xlsx: {exc}") from exc

    try:
        ws = wb[wb.sheetnames[0]]
        header_map: dict[str, int] | None = None
        rows: list[dict[str, Any]] = []
        aliases = {
            "date": {"date", "nav date", "as of date"},
            "nav": {"nav", "net asset value"},
            "shares_outstanding": {"shares outstanding", "sharesoutstanding"},
            "total_net_assets": {"total net assets", "assets under management", "net assets"},
        }

        for cells in ws.iter_rows(values_only=True):
            values = list(cells)
            if header_map is None:
                norm = [_norm_header(v) for v in values]
                candidate: dict[str, int] = {}
                for key, names in aliases.items():
                    for idx, text in enumerate(norm):
                        if text in names:
                            candidate[key] = idx
                            break
                if len(candidate) == 4:
                    header_map = candidate
                continue

            if not any(v not in (None, "") for v in values):
                if rows:
                    break
                continue
            try:
                raw = {key: values[idx] if idx < len(values) else None for key, idx in header_map.items()}
                row = _normalize_nav_row(raw)
            except Exception:
                # Footer/disclaimer or malformed non-data row.
                continue
            rows.append(row)

        if header_map is None:
            raise ValueError("State Street NAV history header not found: Date/NAV/Shares Outstanding/Total Net Assets")
        dedup = {r["date"]: r for r in rows}
        out = sorted(dedup.values(), key=lambda r: r["date"])
        if len(out) < 22:
            raise ValueError(f"State Street NAV history has only {len(out)} valid rows; need at least 22")
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _split_like(prev: dict[str, Any], cur: dict[str, Any]) -> bool:
    try:
        sr = float(cur["shares_outstanding"]) / float(prev["shares_outstanding"])
        nr = float(cur["nav"]) / float(prev["nav"])
        ar = float(cur["total_net_assets"]) / float(prev["total_net_assets"])
    except Exception:
        return False
    # A share split changes share count and NAV inversely while AUM stays roughly
    # unchanged.  Never mislabel that mechanical event as gigantic fund flow.
    return bool(abs(sr - 1.0) >= 0.30 and 0.82 <= sr * nr <= 1.18 and 0.75 <= ar <= 1.25)


def _flow_observations(nav_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for prev, cur in zip(nav_rows, nav_rows[1:]):
        if _split_like(prev, cur):
            continue
        prev_assets = float(prev["total_net_assets"])
        nav = float(cur["nav"])
        share_delta = float(cur["shares_outstanding"]) - float(prev["shares_outstanding"])
        flow_usd = share_delta * nav
        flow_pct = flow_usd / prev_assets * 100.0 if prev_assets > 0 else 0.0
        perf = (nav / float(prev["nav"]) - 1.0) * 100.0
        if not all(math.isfinite(v) for v in (flow_usd, flow_pct, perf)):
            continue
        out.append({
            "date": str(cur["date"]),
            "previous_date": str(prev["date"]),
            "nav": nav,
            "nav_previous": float(prev["nav"]),
            "shares_outstanding": float(cur["shares_outstanding"]),
            "shares_outstanding_previous": float(prev["shares_outstanding"]),
            "shares_delta": share_delta,
            "aum_usd": float(cur["total_net_assets"]),
            "aum_usd_previous": prev_assets,
            "flow_1d_usd": flow_usd,
            "flow_pct": flow_pct,
            "perf_1d_pct": perf,
        })
    if not out:
        raise ValueError("State Street NAV history produced no valid daily flow observations")
    return out


def _strength(today: float, prior: list[dict[str, Any]]) -> dict[str, Any]:
    window = prior[-MIN_HISTORY:]
    if len(window) < MIN_HISTORY:
        return {"available": False, "score": None, "z": None, "sample_prior": len(window), "sample_target": MIN_HISTORY}
    vals = [float(x["flow_pct"]) for x in window]
    mean = statistics.fmean(vals)
    stdev = statistics.stdev(vals)
    if stdev <= 1e-15:
        z = 0.0 if abs(today - mean) <= 1e-15 else (2.5 if today > mean else -2.5)
    else:
        z = (today - mean) / stdev
    score = max(0.0, min(100.0, 50.0 + 20.0 * z))
    return {
        "available": True,
        "score": round(score, 2),
        "z": round(z, 4),
        "sample_prior": len(window),
        "sample_target": MIN_HISTORY,
        "mean_20": round(mean, 8),
        "stdev_20": round(stdev, 8),
    }


def _nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    first = date(year, month, 1)
    delta = (weekday - first.weekday()) % 7
    return first + timedelta(days=delta + 7 * (n - 1))


def _last_weekday(year: int, month: int, weekday: int) -> date:
    last_day = calendar.monthrange(year, month)[1]
    d = date(year, month, last_day)
    return d - timedelta(days=(d.weekday() - weekday) % 7)


def _observed_fixed(d: date) -> date:
    if d.weekday() == 5:
        return d - timedelta(days=1)
    if d.weekday() == 6:
        return d + timedelta(days=1)
    return d


def _easter_sunday(year: int) -> date:
    # Anonymous Gregorian algorithm.
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _nyse_holidays(year: int) -> set[date]:
    return {
        _observed_fixed(date(year, 1, 1)),
        _nth_weekday(year, 1, 0, 3),       # MLK Day
        _nth_weekday(year, 2, 0, 3),       # Washington's Birthday
        _easter_sunday(year) - timedelta(days=2),  # Good Friday
        _last_weekday(year, 5, 0),         # Memorial Day
        _observed_fixed(date(year, 6, 19)),
        _observed_fixed(date(year, 7, 4)),
        _nth_weekday(year, 9, 0, 1),       # Labor Day
        _nth_weekday(year, 11, 3, 4),      # Thanksgiving
        _observed_fixed(date(year, 12, 25)),
    }


def _previous_nyse_session(ref_date: date) -> date:
    d = ref_date - timedelta(days=1)
    while d.weekday() >= 5 or d in _nyse_holidays(d.year):
        d -= timedelta(days=1)
    return d


def _minimum_expected_date() -> date:
    from zoneinfo import ZoneInfo
    ny_today = datetime.now(timezone.utc).astimezone(ZoneInfo("America/New_York")).date()
    return _previous_nyse_session(ny_today)


def _download_nav_history(session: requests.Session, ticker: str, timeout: float) -> tuple[bytes, str]:
    url = NAVHIST_BASE.format(ticker=ticker.lower())
    response = session.get(url, timeout=(8, timeout))
    response.raise_for_status()
    content = response.content
    ctype = str(response.headers.get("content-type") or "").lower()
    if content[:2] != b"PK":
        preview = response.text[:160] if "text" in ctype or "html" in ctype else repr(content[:80])
        raise RuntimeError(
            f"State Street {ticker} NAV history did not return xlsx; "
            f"status={response.status_code}, content-type={ctype!r}, bytes={len(content)}, preview={preview!r}"
        )
    return content, response.url


def _fixture_bytes(fixture_dir: Path, ticker: str) -> bytes:
    for name in (f"{ticker}.xlsx", f"{ticker.lower()}.xlsx", f"navhist-us-en-{ticker.lower()}.xlsx"):
        p = fixture_dir / name
        if p.exists():
            return p.read_bytes()
    raise FileNotFoundError(f"fixture missing for {ticker} in {fixture_dir}")


def collect(*, timeout: float = 25.0, fixture_dir: str | Path | None = None) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    })

    fixture_path = Path(fixture_dir) if fixture_dir else None
    rows: list[dict[str, Any]] = []
    history_out: dict[str, list[dict[str, Any]]] = {}
    errors: list[dict[str, str]] = []

    for meta in SECTORS:
        ticker = meta["ticker"]
        try:
            if fixture_path:
                content = _fixture_bytes(fixture_path, ticker)
                source_url = f"fixture://{ticker}"
                fetch_method = "fixture_xlsx"
            else:
                content, source_url = _download_nav_history(session, ticker, timeout)
                fetch_method = "state_street_navhist_xlsx"

            nav_rows = parse_nav_history_xlsx(content)
            obs = _flow_observations(nav_rows)
            current = obs[-1]
            prior = obs[:-1]
            strength = _strength(float(current["flow_pct"]), prior)
            history_out[meta["id"]] = [
                {"date": x["date"], "flow_pct": round(float(x["flow_pct"]), 8)}
                for x in obs[-HISTORY_KEEP:]
            ]
            row = {
                **meta,
                "source_url": source_url,
                "fetch_method": fetch_method,
                "data_as_of": current["date"],
                "previous_data_as_of": current["previous_date"],
                "data_as_of_source": "state_street_nav_history",
                "nav": round(float(current["nav"]), 8),
                "nav_previous": round(float(current["nav_previous"]), 8),
                "shares_outstanding": round(float(current["shares_outstanding"]), 4),
                "shares_outstanding_previous": round(float(current["shares_outstanding_previous"]), 4),
                "shares_delta": round(float(current["shares_delta"]), 4),
                "flow_1d_usd": round(float(current["flow_1d_usd"]), 2),
                "aum_usd": round(float(current["aum_usd"]), 2),
                "aum_usd_previous": round(float(current["aum_usd_previous"]), 2),
                "flow_pct": round(float(current["flow_pct"]), 8),
                "perf_1d_pct": round(float(current["perf_1d_pct"]), 6),
                "flow_method": "delta_shares_outstanding_x_current_nav",
                "flow_strength": strength,
            }
            rows.append(row)
        except Exception as exc:
            errors.append({"sector": meta["id"], "ticker": ticker, "error": str(exc)})

    if len(rows) != len(SECTORS):
        raise RuntimeError(f"sector flow incomplete: {len(rows)}/{len(SECTORS)}; errors={errors}")

    data_dates = sorted({str(x["data_as_of"]) for x in rows})
    if len(data_dates) != 1:
        raise RuntimeError(f"State Street sector dates are mixed; refusing HUD publish: {data_dates}")

    source_date = datetime.strptime(data_dates[0], "%Y-%m-%d").date()
    expected_min = _minimum_expected_date()
    if not fixture_path and source_date < expected_min:
        raise RuntimeError(
            f"State Street NAV history is stale: source={source_date.isoformat()}, "
            f"minimum_expected={expected_min.isoformat()}; refusing to overwrite R2 with old flow data"
        )

    positive = [x for x in rows if float(x["flow_pct"]) > 0]
    pool = positive or rows
    all_strength_ready = all(bool(x.get("flow_strength", {}).get("available")) for x in rows)
    if all_strength_ready:
        leader = max(pool, key=lambda x: (float(x["flow_strength"]["score"]), float(x["flow_pct"])))
        leader_method = "flow_strength_20d_state_street"
    else:
        leader = max(pool, key=lambda x: float(x["flow_pct"]))
        leader_method = "flow_pct_until_20d_ready"

    now = datetime.now(TW)
    return {
        "schema_version": "sector-flow-v2-state-street-spdr-navhist",
        "generated_at_taiwan": now.strftime("%Y-%m-%d %H:%M:%S"),
        "source": "State Street Select Sector SPDR NAV History",
        "schedule": "21:31 Asia/Taipei",
        "metric": "Estimated 1D primary-market flow = Δ Shares Outstanding × NAV; normalized by prior AUM",
        "strength_formula": "clamp(50 + 20*zscore(today_flow_pct vs previous 20 valid sessions), 0, 100)",
        "minimum_expected_data_date": expected_min.isoformat(),
        "data_dates": data_dates,
        "leader": leader["id"],
        "leader_method": leader_method,
        "flow_regime": "positive_inflow_exists" if positive else "all_non_positive",
        "sectors": rows,
        "history": history_out,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output")
    ap.add_argument("--previous", help="kept for workflow backward compatibility; State Street navhist now supplies history directly")
    ap.add_argument("--fixture", help="offline parser test: parse one State Street navhist xlsx")
    ap.add_argument("--fixture-dir", help="offline end-to-end test directory containing XLK/XLF/... xlsx fixtures")
    args = ap.parse_args()

    if args.fixture:
        nav_rows = parse_nav_history_xlsx(Path(args.fixture).read_bytes())
        obs = _flow_observations(nav_rows)
        print(json.dumps({"rows": len(nav_rows), "latest": nav_rows[-1], "latest_flow": obs[-1]}, ensure_ascii=False, indent=2))
        return 0

    if not args.output:
        ap.error("--output is required unless --fixture is used")
    result = collect(fixture_dir=args.fixture_dir)
    target = Path(args.output)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "ok": True,
        "source": result["source"],
        "leader": result["leader"],
        "leader_method": result["leader_method"],
        "dates": result["data_dates"],
        "minimum_expected_data_date": result["minimum_expected_data_date"],
        "sectors": len(result["sectors"]),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
